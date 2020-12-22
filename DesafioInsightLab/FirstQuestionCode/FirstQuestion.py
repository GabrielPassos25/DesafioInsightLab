#Bibliotecas utilizadas: [openpyxl, pathlib, string, re, warnings]
import openpyxl #Utilizada para abrir e ler arquivos .xlxs
import string #Facilitamento do tratamento das strings
from pathlib import Path #Pré-requisito da biblioteca openpyxl
import re #Tratamento de links, acentuação e pontuação
import warnings 
#Ignorar o warning que dizia que a biblioteca openpyxl não possui uma versão "default", somente por convenção
warnings.filterwarnings("ignore")

#Leitura dos arquivos
def input_data(d):
    #Inicialização dataset1
    dataset1_path = Path('','Data/RelatorioDePublicacoes-_01_Abr_2020a09_Abr_2020.xlsx')
    dataset1_file = openpyxl.load_workbook(dataset1_path)
    dataset1 = dataset1_file.active
    
    #Contagem de palavras dataset1
    d = count_words(dataset1,d)

    #Inicialização dataset2
    dataset2_path = Path('','Data/RelatorioDePublicacoes-_09_Abr_2020a16_Abr_2020.xlsx')
    dataset2_file = openpyxl.load_workbook(dataset2_path)
    dataset2 = dataset2_file.active

    #Contagem de palavras dataset2
    d = count_words(dataset2,d)
    return d

#Escrita resposta
def archive_write(d):
    #Abertura do arquivo
    archive = open("AnswerFirstQuestion.txt","w")
    #Escrita do arquivo, já com a contagem
    for i in list(d.keys()):
        archive.write(str(i) + ": " + str(d[i]) + "\n")
    #Fechamento do arquivo (Resposta final)
    archive.close()

#Contagem de palavras
def count_words(dataset,d):
    #Checando se existe link no comentário, para evitar a formatação do mesmo
    regex = r"(?i)\b((?:https?://|www\d{0,3}[.]|[a-z0-9.\-]+[.][a-z]{2,4}/)(?:[^\s()<>]+|\(([^\s()<>]+|(\([^\s()<>]+\)))*\))+(?:\(([^\s()<>]+|(\([^\s()<>]+\)))*\)|[^\s`!()\[\]{};:'\".,<>?«»“”‘’]))"
    #Fluxo básico de contagem e tratamento de string
    for i, row in enumerate(dataset.iter_rows(values_only=True)):
        #Ignorar cabeçalho
        if i >0:
            #Coluna conteúdo
            linha = row[11]
            #Retirando espaços e quebra de linha, além de deixar tudo em letra minúscula
            linha = re.sub("[\\(\\)\\[\\]\\{\\}]","",linha)
            linha = linha.lower()
            #Separando Links
            url = re.findall(regex,linha)
            list_url = [x[0] for x in url]
            #Quebrando as palavras em espaços
            palavras = linha.split()
            #Adicionando Links
            for palavra in palavras:
                if palavra in list_url:
                    if palavra in d:
                        d[palavra] = d[palavra] + 1
                    else:
                        d[palavra] = 1
                    continue
                #Tratando acentuação
                palavra = re.sub("[áàäâã]", "a", palavra)
                palavra = re.sub("[éèëê]", "e", palavra)
                palavra = re.sub("[íìïî]", "i", palavra)
                palavra = re.sub("[óòöôõ]", "o", palavra)
                palavra = re.sub("[úùüû]", "u", palavra)
                palavra = re.sub("[ýÿ]", "y", palavra)
                #Tratando pontuação
                palavra = re.sub("[^a-z^A-Z^0-9]"," ", palavra)
                #Adicionando palavras
                for palavra_format in palavra.split(): 
                    if palavra_format in d:
                        d[palavra_format] = d[palavra_format] + 1
                    else:
                        d[palavra_format] = 1
    d = dict(sorted(d.items()))
    return d

#Fluxo do programa
def main():
    global d
    d = dict()
    #Saída para terminal
    print("Realizando leitura dos dados...")
    #Execução das funções
    d = input_data(d)
    archive_write(d)
    #Saída final do programa
    print('\033[92m' + "Leitura realizada!" + '\033[92m')
#Chamada função main
main()